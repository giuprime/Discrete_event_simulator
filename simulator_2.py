
import heapq
"""
Heapq is module that provides an implementation of the heap queue algorithm. In this algorithm there're binary trees for
which every parent node has a value less than or equal to any of its children. The property of this data structure is 
that each we pop or push an element the heap structure is mantained
"""
import random
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active



# definisco la classe Evento
class Event:
    def __init__(self, time, type):
        # Descrive il tempo (random) dell'arrivo di un nuovo cliente
        self.time = time
        # Descrive il tipo di evento "arrival" o "departure"
        self.type = type

    def __lt__(self, other):
        # Questo metodo (less than) che fa parte della classe Event viene utilizzato per il confronto minore < tra due oggetti di
        # tipo event. In questo caso viene usata per confrontare i 2 oggetti 'self' e 'other' in base al loro attributo
        # time. Questo confronto è necessario per estarre in ordine cronologico gli eventi della coda
        return self.time < other.time


class Customer:
    def __init__(self, arrival_time):
        # Definizione degli attributi
        self.arrival_time = arrival_time
        self.service_start_time = 0
        self.departure_time = 0
        self.wait_time = 0


class Simulator:
    def __init__(self, arrival_rate, service_rate):
        # Definizione degli attributi
        self.current_time = 0
        # dichiaro la lista degli eventi
        self.event_queue = []
        # serve solo per stampare gli eventi ma non ha un ruolo nel simulatore
        self.event_list = []
        self.server_busy = False
        self.customers_in_queue = []
        self.total_wait_time = 0
        self.total_service_time = 0
        self.customers_served = 0
        self.customer_counter = 0
        self.customers_in_system = []

        self.arrival_rate = arrival_rate
        self.service_rate = service_rate

    def schedule_event(self, event):
        # Inserimento nella coda degli eventi tenendo in considerazione l'ordine minimo degli eventi. Viene usato unheap
        # binario per mantenere gli eventi con l'ordine determinato dal tempo minimo.
        # In questo caso l'evento 'event' viene inserito nella coda degli eventi, così facendo il tempo minore si trova
        # sempre nella radice dell'heap. Grazie a ciò gli eventi vengono elaborati con il giusto ordine
        heapq.heappush(self.event_queue, event)
        self.event_list.append(event)

    def generate_interarrival_time(self):
        # expovariate ritorna un numero float random seguendo la distribuzione esponenziale
        return random.expovariate(self.arrival_rate)

    def generate_service_time(self):
        return random.expovariate(self.service_rate)

    def handle_arrival_event(self):
        self.customer_counter += 1
        print("Arrivo del cliente", self.customer_counter ,"al tempo:", round(self.current_time, 2), "s")

        # Creo un'istanza della classe Customer
        customer = Customer(self.current_time)

        # se il server è occupato
        if self.server_busy:
            # aggiunge il cliente nella lista dei clienti
            self.customers_in_queue.append(customer)

        else:
            # cambio lo stato del server
            self.server_busy = True
            # genero un tempo di servizio
            service_time = self.generate_service_time()

            departure_time = self.current_time + service_time
            # dopo aver fatto i calcoli sopra richiamo il metodo di cui sotto che genera un nuovo evento "departure"
            # perché il cliente parte ed esce dal sistema
            departure_event = Event(departure_time, "departure")
            # viene programmato l'evento di partenza e aggiunto alla lista degli eventi
            self.schedule_event(departure_event)
            # aggiornamento del tempo totale del servizio
            self.total_service_time += service_time

            # imposto il tempo di servizio e di partenza per il cliente corrente
            customer.service_start_time = self.current_time
            customer.departure_time = departure_time
        # generazione del tempo che passa tra un cliente ed un altro
        interarrival_time = self.generate_interarrival_time()
        # calcolo del tempo di arrivo
        arrival_time = self.current_time + interarrival_time
        # creazione di un nuovo evento di tipo "arrival"
        arrival_event = Event(arrival_time, "arrival")
        # viene aggiunto il nuovo evento alla coda degli eventi
        self.schedule_event(arrival_event)

        self.current_time = arrival_time

    def handle_departure_event(self):
        self.customer_counter -= 1
        print("Partenza del cliente",self.customer_counter, "al tempo:", round(self.current_time, 2), "s")
        # controllo se ci sono clienti in coda
        if self.customers_in_queue:
            # viene estratto il primo cliente nella coda e passato alla variabile
            next_customer = self.customers_in_queue.pop(0)
            service_time = self.generate_service_time()
            departure_time = self.current_time + service_time
            if departure_time <= 60:
                self.customers_served += 1
            # viene generato un evento "departure"
            departure_event = Event(departure_time, "departure")
            self.schedule_event(departure_event)
            self.total_service_time += service_time

            #vengono impostati il tempo di inizio servizio e di partenza per il nuovo cliente
            next_customer.service_start_time = self.current_time
            next_customer.departure_time = departure_time
            # calcolo
            next_customer.wait_time = self.current_time - next_customer.arrival_time

            # il tempo di attesa del cliente viene aggiunto al tempo totale di attesa
            self.total_wait_time += next_customer.wait_time

            print("Inizio servizio al tempo:", round(self.current_time), "s")
            print("Partenza del cliente al tempo:", round(departure_time), "s")
        else:

            self.server_busy = False
            print("Il server è libero.")

    def run(self, end_time):
        # viene generato il tempo di arrivo
        arrival_time = self.generate_interarrival_time()
        # richiamo la classe Event e le passo il tempo di arrivo e il parametro arrival per generare un evento arrival
        arrival_event = Event(arrival_time, "arrival")
        # viene programmato l'evento di partenza e aggiunto alla lista degli eventi
        self.schedule_event(arrival_event)
        # controllo se c'è un evento nella coda e se il tempo corrente è minore del tempo finale
        while self.event_queue and self.current_time < end_time:
            # estraggo un evento dalla lista degli eventi
            current_event = heapq.heappop(self.event_queue)
            # assegno il tempo corrente
            self.current_time = current_event.time
            # controllo il tipo di evento estratto
            if current_event.type == "arrival":
                self.handle_arrival_event()
            elif current_event.type == "departure":
                self.handle_departure_event()
        print("FINE SIMULAZIONE")



    def print_event_list(self):
        print("\nLista degli eventi: ")
        for event in self.event_list:
           print(f"Tempo: {event.time}, Tipo: {event.type}")

    def calculate_served_customers_1m(self):
        return simulator.customers_served

    def calculate_troughput(self):
        customer_served = self.calculate_served_customers_1m()
        troughput = customer_served / simulation_time
        return troughput


    def calculate_total_customers(self):
        total_customers = len(self.customers_in_queue) + int(self.server_busy)
        return total_customers

    def calculate_average_wait_time(self):
        total_customers = self.calculate_total_customers()
        if total_customers > 0:
            average_wait_time = self.total_wait_time / total_customers
        else:
            return 0
        return average_wait_time

    def calculate_average_service_time(self):
        total_customers = self.calculate_total_customers()
        if total_customers > 0:
            average_service_time = self.total_service_time / total_customers
        else:
            return 0
        return average_service_time

    def calculate_server_utilization(self):

        server_utilization = (simulator.total_service_time/simulator.current_time)
        return server_utilization

    def calculate_expected_queue_lenght(self):
        total_wait_time = sum(customer.wait_time for customer in self.customers_in_queue)
        total_service_time = self.total_service_time
        current_time = self.current_time

        if current_time > 0:
            expected_queue_lenght = (total_wait_time + total_service_time) / current_time
        else:
            return 0
        return expected_queue_lenght

#-----------------------------------------------------------------------------------------------------------------------
    def print_simulation_statistics(self):
        print("\n**Statistiche simulazione**")

        total_customers = self.calculate_total_customers()
        print("Numero totale dei clienti: ", total_customers)

        average_wait_time= self.calculate_average_wait_time()
        print("Tempo medio di attesa: ", round(average_wait_time, 2), "s")

        average_service_time = self.calculate_average_service_time()
        print("Tempo medio di servizio: ", round(average_service_time, 2), "s")

        server_utilization = self.calculate_server_utilization()
        print("Percentuale di utilizzo del server: ", round(server_utilization, 2), "%")

        expected_queue_lenght = self.calculate_expected_queue_lenght()
        print("Lunghezza prevista della coda: ", round(expected_queue_lenght, 2))

        customers_served_1m = self.calculate_served_customers_1m()
        print("Il numero di clienti serviti in un minuto è: ", customers_served_1m)

        troughput = self.calculate_troughput()
        print("Il troughput equivale a:", troughput, "clienti al minuto")


# ----------------------------------------------------------------------------------------------------------------------
#intestazioni delle colonne del file Excel
sheet['A1'] = "Numero totale dei clienti"
sheet['B1'] = "Tempo medio di attesa"
sheet['C1'] = "Tempo medio di servizio"
sheet['D1'] = "Percentuale di utilizzo del server"
sheet['E1'] = "Lunghezza prevista della coda"

row = 2

arrival_rate = float(input("Inserire il valore dell'arrival rate: "))   #0.5
service_rate = float(input("Inserire il valore del service rate: "))    #0.2
simulation_time = int(input("Inserire il simulation time: "))           #100
num_ripetizione = int(input("Inserire quante volte vuoi ripetere la simulazione: "))

for x in range(num_ripetizione):
    simulator = Simulator(arrival_rate, service_rate)
    simulator.run(simulation_time)
    simulator.print_simulation_statistics()
    #simulator.print_event_list()

    # scrittura delle statistiche all'interno del file Excel
    sheet.cell(row=row, column=1).value = simulator.calculate_total_customers()
    sheet.cell(row=row, column=2).value = round(simulator.calculate_average_wait_time(), 2)
    sheet.cell(row=row, column=3).value = round(simulator.calculate_average_service_time(), 2)
    sheet.cell(row=row, column=4).value = round(simulator.calculate_server_utilization(), 2)
    sheet.cell(row=row, column=5).value = round(simulator.calculate_expected_queue_lenght(), 2)

    row +=1

workbook.save("statistics_2.xlsx")






